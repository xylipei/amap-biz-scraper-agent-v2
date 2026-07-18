"""
文件导出模块 - 将结构化数据导出为CSV或Excel

职责：
- 将结构化数据列表导出为CSV（utf-8-sig）或Excel（.xlsx）
- 文件名包含区域与品类信息
- 返回文件保存的绝对路径

防跑偏要求（PRD 4.1-4）：
- CSV需使用 utf-8-sig 编码处理中文
- 必须返回文件保存的绝对路径
"""

import csv
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# 输出目录
OUTPUT_DIR = "output"


def _ensure_output_dir() -> str:
    """确保输出目录存在，返回绝对路径"""
    abs_dir = os.path.abspath(OUTPUT_DIR)
    os.makedirs(abs_dir, exist_ok=True)
    return abs_dir


def _build_filename(
    region: str,
    keyword: str,
    extension: str = ".csv",
) -> str:
    """
    构建文件名：{区域}_{品类}_{YYYYMMDD}.{ext}

    参数：
        region: 区域名
        keyword: 品类关键词
        extension: 文件扩展名（.csv 或 .xlsx）
    """
    today = datetime.now().strftime("%Y%m%d")
    # 清理文件名中的非法字符
    safe_region = "".join(c for c in region if c.isalnum() or c in " ()-_")
    safe_keyword = "".join(c for c in keyword if c.isalnum() or c in " ()-_")
    return f"{safe_region}_{safe_keyword}_{today}{extension}"


def _get_export_fields() -> List[str]:
    """获取导出字段列表（统一顺序）"""
    return [
        "name",
        "address",
        "location",
        "pname",
        "cityname",
        "adname",
        "tel",
        "type",
        "same_name_count",
        "groupbuy",
        "groupbuy_url",
        "collect_year",
    ]


def _sanitize_csv_cell(value: str) -> str:
    """
    清洗 CSV 单元格值，防止公式注入。

    如果值以 = + - @ 开头，在前面加单引号，
    使 Excel/WPS 将其视为文本而非公式。
    """
    if value and value[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value


def _prepare_row(item: Dict[str, Any], fields: List[str]) -> Dict[str, Any]:
    """准备导出行：确保所有字段存在，处理中文和特殊字符"""
    row = {}
    for field in fields:
        value = item.get(field, "")
        if value is None:
            value = ""
        # 确保字符串类型
        if not isinstance(value, str):
            value = str(value)
        row[field] = _sanitize_csv_cell(value)
    return row


def export_to_table(
    data: List[Dict[str, Any]],
    filename: Optional[str] = None,
    region: str = "unknown",
    keyword: str = "unknown",
    fmt: str = "csv",
) -> str:
    """
    将结构化数据导出为表格文件。

    参数：
        data: 结构化数据列表
        filename: 文件名（可选，不指定则自动生成）
        region: 区域名（用于自动生成文件名）
        keyword: 品类关键词（用于自动生成文件名）
        fmt: 输出格式，"csv" 或 "xlsx"

    返回：
        文件保存的绝对路径

    抛出：
        RuntimeError: 导出失败时
    """
    if not data:
        logger.warning("无数据可导出，跳过文件生成")
        raise RuntimeError("无数据可导出")

    output_dir = _ensure_output_dir()
    fields = _get_export_fields()

    if not filename:
        ext = ".xlsx" if fmt == "xlsx" else ".csv"
        filename = _build_filename(region, keyword, extension=ext)

    # 确保文件名在输出目录中
    filepath = os.path.join(output_dir, filename)
    abs_path = os.path.abspath(filepath)

    try:
        if fmt == "xlsx":
            abs_path = _export_excel(data, fields, abs_path)
        else:
            abs_path = _export_csv(data, fields, abs_path)

        logger.info("文件已保存: %s (共 %d 行)", abs_path, len(data))
        print(f"[完成] 数据已导出到: {abs_path}")
        return abs_path

    except Exception as e:
        error_msg = f"文件导出失败: {e}"
        logger.error(error_msg)
        raise RuntimeError(error_msg) from e


def _export_csv(
    data: List[Dict[str, Any]],
    fields: List[str],
    filepath: str,
) -> str:
    """
    导出为CSV，使用 utf-8-sig 编码以兼容 Excel 中文显示。

    参数：
        data: 数据列表
        fields: 字段列表
        filepath: 文件路径（含扩展名）

    返回：
        文件绝对路径
    """
    if not filepath.endswith(".csv"):
        filepath += ".csv"

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for item in data:
            writer.writerow(_prepare_row(item, fields))

    return os.path.abspath(filepath)


def _export_excel(
    data: List[Dict[str, Any]],
    fields: List[str],
    filepath: str,
) -> str:
    """
    导出为Excel (.xlsx)，使用 openpyxl。

    参数：
        data: 数据列表
        fields: 字段列表
        filepath: 文件路径（含扩展名）

    返回：
        文件绝对路径
    """
    if not filepath.endswith(".xlsx"):
        filepath += ".xlsx"

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "商家数据"

    # 写表头
    header_row = []
    field_display = {
        "name": "门店名称",
        "address": "地址",
        "location": "经纬度",
        "pname": "省份",
        "cityname": "城市",
        "adname": "区县",
        "tel": "电话",
        "type": "POI类型",
        "same_name_count": "同名门店数量",
        "groupbuy": "是否支持团购",
        "groupbuy_url": "团购详情页",
        "collect_year": "高德收录年份",
    }
    for field in fields:
        header_row.append(field_display.get(field, field))
    ws.append(header_row)

    # 写数据行
    for item in data:
        row = []
        for field in fields:
            value = item.get(field, "")
            if value is None:
                value = ""
            if not isinstance(value, str):
                value = str(value)
            row.append(value)
        ws.append(row)

    # 自动调整列宽
    for col_idx in range(1, len(fields) + 1):
        max_length = len(str(header_row[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            cell_value = row[0]
            if cell_value:
                max_length = max(max_length, min(len(str(cell_value)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    wb.save(filepath)
    return os.path.abspath(filepath)


# ── 搜索记录 ──

HISTORY_FIELDS = [
    "search_time",
    "user_input",
    "region",
    "keyword",
    "modifier",
    "total",
    "groupbuy_yes",
    "groupbuy_failed",
    "file_path",
]


def save_search_history(metadata: Dict[str, Any]) -> str:
    """
    将一次搜索的元信息追加写入 search_history.csv。

    参数：
        metadata: 字典，至少包含 HISTORY_FIELDS 中定义的键

    返回：
        搜索记录文件的绝对路径
    """
    output_dir = _ensure_output_dir()
    history_file = os.path.join(output_dir, "search_history.csv")
    file_exists = os.path.exists(history_file)

    with open(history_file, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=HISTORY_FIELDS)
        if not file_exists:
            writer.writeheader()
        row = {k: metadata.get(k, "") for k in HISTORY_FIELDS}
        # 对字符串字段做公式注入防护
        for k, v in row.items():
            if isinstance(v, str):
                row[k] = _sanitize_csv_cell(v)
        writer.writerow(row)

    logger.info("搜索记录已保存 -> %s", history_file)
    return os.path.abspath(history_file)


def get_search_history_path() -> str:
    """返回搜索记录文件的绝对路径"""
    output_dir = _ensure_output_dir()
    return os.path.join(output_dir, "search_history.csv")
