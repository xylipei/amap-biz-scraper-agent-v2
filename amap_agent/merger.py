"""
跨文件合并去重 + 品牌聚合模块

职责：
- 读取 output/ 下多个商家 CSV（utf-8-sig，表头为中文）
- 按 (门店名称, 地址) 联合去重（防止同名不同址的分店被误删）
- 品牌识别：取店铺名称括号前文本作为品牌名（与 aggregator._extract_base_name 规则一致）
- 全局重算「同名门店数量」，按品牌门店数降序排序
- 导出 Excel 品牌分组明细表（品牌名/门店数合并单元格 + 品牌序号）

供 merge_brands.py CLI 与 merge-center-results skill 调用。
"""

import csv
import logging
import os
import re
from collections import Counter
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# 与 exporter._FIELD_DISPLAY 反向映射：中文表头 -> 内部字段
_HEADER_TO_FIELD = {
    "门店名称": "name",
    "同名门店数量": "same_name_count",
    "高德收录年份": "collect_year",
    "省份": "pname",
    "城市": "cityname",
    "区县": "adname",
    "地址": "address",
    "电话": "tel",
    "评分": "rating",
    "POI类型": "type",
    "是否团购": "groupbuy",
    "团购链接": "groupbuy_url",
}

# 内部字段 -> Excel 明细表头（品牌分组明细表用，与源 CSV 列顺序解耦）
_BRAND_TABLE_HEADERS = [
    "序号",
    "品牌名",
    "同名门店数量",
    "门店名称",
    "地址",
    "电话",
    "评分",
    "是否团购",
    "团购链接",
]

# 明细列：源字段 key -> 内部字段
_BRAND_DETAIL_FIELDS = [
    "name",
    "address",
    "tel",
    "rating",
    "groupbuy",
    "groupbuy_url",
]

_DEFAULT_OUTPUT_NAME = "merged_brands_{date}.xlsx"


def _strip_formula_prefix(value: str) -> str:
    """
    去掉 CSV 导出时的公式注入防护前缀（exporter._sanitize_csv_cell 会给
    以 = + - @ 开头的值加前导单引号），读取时还原为原始值。
    """
    if value.startswith("'") and len(value) > 1 and value[1] in ("=", "+", "-", "@", "\t", "\r"):
        return value[1:]
    return value


def load_csv_files(paths: List[str]) -> List[Dict[str, Any]]:
    """
    读取多个商家 CSV，统一字段结构。

    参数：
        paths: CSV 文件路径列表

    返回：
        记录字典列表，字段为内部英文 key（name/address/tel/rating/...），
        另附 _source_file 标记来源文件。

    说明：
        - 编码优先 utf-8-sig（兼容 Excel 导出的中文 BOM），失败时回退 gb18030
        - 未知列忽略；文件缺少「门店名称」或「地址」列时告警并跳过该文件
          （地址列是去重键组成部分，缺失会导致同名不同址分店被误合并）
    """
    records: List[Dict[str, Any]] = []
    for path in paths:
        if not os.path.exists(path):
            logger.warning("文件不存在，跳过: %s", path)
            continue
        data = _read_csv_with_fallback(path)
        if data is None:
            logger.warning("文件编码无法识别（utf-8/gb18030 均失败），跳过: %s", path)
            continue
        reader, fieldnames = data
        if not fieldnames:
            logger.warning("空文件，跳过: %s", path)
            continue
        # 表头 -> 内部字段映射（未识别的中文表头列丢弃）
        col_map = {}
        for header in fieldnames:
            header = (header or "").strip()
            if header in _HEADER_TO_FIELD:
                col_map[header] = _HEADER_TO_FIELD[header]
        if "name" not in col_map.values():
            logger.warning("文件缺少「门店名称」列，跳过: %s", path)
            continue
        if "address" not in col_map.values():
            logger.warning("文件缺少「地址」列（去重键组成部分），跳过: %s", path)
            continue
        file_count = 0
        for row in reader:
            item: Dict[str, Any] = {"_source_file": os.path.basename(path)}
            for header, field in col_map.items():
                item[field] = _strip_formula_prefix((row.get(header) or "").strip())
            records.append(item)
            file_count += 1
        logger.info("已读取 %s: %d 条", os.path.basename(path), file_count)
    return records


def _read_csv_with_fallback(path: str):
    """
    读取 CSV 并返回 (reader, fieldnames)；utf-8-sig 解码失败时回退 gb18030
    （兼容中文 Excel 导出的 GBK 编码文件）。

    返回：
        (csv.DictReader, fieldnames)；读取失败返回 None
    """
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            f = open(path, "r", newline="", encoding=encoding)
            reader = csv.DictReader(f)
            return reader, reader.fieldnames
        except UnicodeDecodeError:
            continue
        except (csv.Error, OSError) as e:
            logger.warning("读取文件失败: %s (%s)", path, e)
            return None
    return None


def merge_dedupe(records: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], int]:
    """
    按 (门店名称, 地址) 联合去重，保留首见记录。

    参数：
        records: load_csv_files 返回的记录列表

    返回：
        (去重后的记录列表, 被去除的重复条数)
    """
    seen = set()
    deduped: List[Dict[str, Any]] = []
    removed = 0
    for rec in records:
        key = (rec.get("name", ""), rec.get("address", ""))
        if key in seen:
            removed += 1
            continue
        seen.add(key)
        deduped.append(rec)
    if removed:
        logger.info("去重: %d -> %d 条（去除重复 %d 条）", len(records), len(deduped), removed)
    return deduped, removed


def extract_brand(name: str) -> str:
    """
    从完整店名提取品牌名：括号前文本。

    "鲜丰水果(上海荣顺苑店)" -> "鲜丰水果"
    "百果园（莘建路店）"     -> "百果园"
    "星巴克"                 -> "星巴克"（无括号不处理）

    规则与 aggregator._extract_base_name 保持一致。
    """
    if not name:
        return ""
    match = re.match(r"^(.+?)[（(]", name)
    if match:
        return match.group(1).strip()
    return name.strip()


def aggregate_by_brand(
    records: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    按品牌聚合：重算同名门店数量（去重后的全局统计），品牌按门店数降序。

    参数：
        records: 去重后的记录列表

    返回：
        记录列表（保持原记录），每项附加字段：
        - brand: 品牌名
        - brand_rank: 品牌序号（按门店数降序，从 1 开始）
        - same_name_count: 全局重算的同名门店数量（覆盖源文件局部值）
    """
    if not records:
        return []

    for rec in records:
        rec["brand"] = extract_brand(rec.get("name", ""))

    # 过滤无品牌名的记录（空门店名），避免产生 rank=0 的伪品牌块与空品牌计数
    before = len(records)
    records = [r for r in records if r["brand"]]
    if len(records) != before:
        logger.warning("已过滤无品牌名的记录 %d 条", before - len(records))
        if not records:
            return []

    # 全局品牌门店数
    brand_counter = Counter(rec["brand"] for rec in records if rec["brand"])

    # 品牌按门店数降序（并列时按品牌名字典序稳定排序）
    brands_sorted = sorted(
        brand_counter.keys(),
        key=lambda b: (-brand_counter[b], b),
    )
    brand_rank = {b: i + 1 for i, b in enumerate(brands_sorted)}

    for rec in records:
        brand = rec.get("brand", "")
        rec["same_name_count"] = brand_counter.get(brand, 0)
        rec["brand_rank"] = brand_rank.get(brand, 0)

    # 块内按品牌排名（门店数降序）排列，块内按门店名称排序
    records.sort(key=lambda r: (r.get("brand_rank", 0), r.get("brand", ""), r.get("name", "")))

    logger.info(
        "品牌聚合完成: %d 条记录 -> %d 个品牌",
        len(records),
        len(brands_sorted),
    )
    return records


def _style_header(ws, col_count: int) -> None:
    """表头加粗 + 灰底 + 居中 + 边框"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    fill = PatternFill("solid", fgColor="D9E1F2")
    font = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def export_brand_grouped_excel(
    records: List[Dict[str, Any]],
    output_path: str,
) -> str:
    """
    导出品牌分组明细表 Excel。

    布局（品牌分组明细表）：
        - 第 1 列：品牌序号（品牌排名）
        - 第 2 列：品牌名（同品牌行垂直合并单元格）
        - 第 3 列：同名门店数量（同品牌行垂直合并单元格）
        - 第 4 列起：门店明细（门店名称/地址/电话/评分/是否团购/团购链接）

    参数：
        records: aggregate_by_brand 后的记录列表
        output_path: 输出 .xlsx 路径

    返回：
        输出文件绝对路径
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    if not records:
        raise RuntimeError("无数据可导出")

    def _safe_cell_value(value):
        """防公式注入：以 = + - @ 开头的值加前导单引号，避免 openpyxl 将字符串解析为公式。"""
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@", "\t", "\r")):
            return "'" + value
        return value

    wb = Workbook()
    ws = wb.active
    ws.title = "品牌汇总明细"

    header_row = _BRAND_TABLE_HEADERS
    ws.append(header_row)
    _style_header(ws, len(header_row))

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # 按品牌分组（aggregate_by_brand 已按 brand 排序）
    row_idx = 2
    i = 0
    while i < len(records):
        rec = records[i]
        brand = rec.get("brand", "")
        # 统计本品牌连续行数
        j = i
        while j < len(records) and records[j].get("brand", "") == brand:
            j += 1
        block_len = j - i

        # 品牌块首行写 序号/品牌名/门店数，随后合并单元格
        start_row = row_idx
        ws.cell(row=row_idx, column=1, value=rec.get("brand_rank", ""))
        ws.cell(row=row_idx, column=2, value=brand)
        ws.cell(row=row_idx, column=3, value=rec.get("same_name_count", 0))
        for col in (1, 2, 3):
            if block_len > 1:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=start_row + block_len - 1,
                    end_column=col,
                )
            ws.cell(row=start_row, column=col).alignment = center
            ws.cell(row=start_row, column=col).border = border

        # 明细列逐店写
        for k in range(i, j):
            detail = records[k]
            col = 4
            for field in _BRAND_DETAIL_FIELDS:
                cell = ws.cell(row=row_idx, column=col, value=_safe_cell_value(detail.get(field, "")))
                cell.border = border
                cell.alignment = left if field in ("name", "address", "groupbuy_url") else center
                col += 1
            row_idx += 1
        i = j

    # 列宽自适应
    for col_idx in range(1, len(header_row) + 1):
        max_len = len(str(header_row[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_value in row:
                if cell_value is not None:
                    max_len = max(max_len, min(len(str(cell_value)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(output_path)
    logger.info("Excel 已保存: %s (%d 行)", output_path, len(records))
    return os.path.abspath(output_path)


def run_merge(
    input_paths: List[str],
    output_dir: str = "output",
    output_path: Optional[str] = None,
) -> Dict[str, Any]:
    """
    一键合并入口：读取 -> 去重 -> 品牌聚合 -> Excel 导出。

    参数：
        input_paths: 输入 CSV 文件列表
        output_dir: 输出目录（自动创建）
        output_path: 输出文件路径（不指定则用 output_dir/merged_brands_{YYYYMMDD}.xlsx）

    返回：
        结果字典：success / stats / file_path / message
    """
    records = load_csv_files(input_paths)
    if not records:
        return {"success": False, "error": "未从输入文件中读取到任何商家数据"}

    merged, removed = merge_dedupe(records)
    if not merged:
        return {"success": False, "error": "去重后无数据可导出"}

    grouped = aggregate_by_brand(merged)

    os.makedirs(output_dir, exist_ok=True)
    if not output_path:
        output_path = os.path.join(
            output_dir, _DEFAULT_OUTPUT_NAME.format(date=datetime.now().strftime("%Y%m%d"))
        )
    if not output_path.endswith(".xlsx"):
        output_path += ".xlsx"
    # 确保输出文件父目录存在（显式指定 output_path 时可能不在 output_dir 下）
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    file_path = export_brand_grouped_excel(grouped, output_path)

    brands = len({r.get("brand", "") for r in grouped})
    stats = {
        "source_files": len(input_paths),
        "total_before_dedupe": len(records),
        "removed_duplicates": removed,
        "total_after_dedupe": len(merged),
        "brand_count": brands,
    }
    return {
        "success": True,
        "stats": stats,
        "file_path": file_path,
        "message": (
            f"合并完成：读取 {len(input_paths)} 个文件，去重前 {len(records)} 条，"
            f"去除重复 {removed} 条，去重后 {len(merged)} 条，共 {brands} 个品牌，"
            f"文件保存至: {file_path}"
        ),
    }
